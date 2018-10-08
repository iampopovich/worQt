#!/usr/bin/pythom
#-*-coding: utf-8-*-
import openpyxl
import time as tt
import datetime as dt
import re
import os
import sys
import lib_headers as hh #можно будет отказаться 

def whatOSRun(path):
	if os.name.endswith('nt'):
		return '%s\\' %path
	else: return '%s/' %path

def findFiles(cwd):
	currentYear = dt.datetime.now().year
	files = [os.path.join(cwd, f) for f in os.listdir(cwd) if os.path.isfile(os.path.join(cwd, f))]
	try:
		cJournalName = '%sСводный_журнал_ДТЭ_%s.xlsx' %(cwd,currentYear)
		if cJournalName in files: files.pop(files.index(cJournalName))
		#else: createJournal(cJournalName) #pass , потому что потом все равно создадим его в wb.asve()
		xlFiles = [item for item in files if not(re.search(r'.*\sДТЭ\s.*%s.*xlsm' %currentYear, item) is None) and not('$' in item)]
		return xlFiles,cJournalName
	except Exception as ex:
		print('findFiles failed with: %s' %ex)
		return None, None

def createJournal(name = None):
	try:
		wb = openpyxl.Workbook()
		wb.remove(wb.active) #remove default worksheet
		for sheetName in ['ДД','ДТЭ']:
			wb.create_sheet(sheetName)
			activeSheet = wb[sheetName] #можно будет отказаться
			activeSheet.append(hh.getHeaders(sheetName)) #можно будет отказаться
			for item in activeSheet.columns:
				activeSheet.column_dimensions['%s' %item[0].column].width = 16.0
			for cell in list(activeSheet.rows)[0]:
				cell.style = 'Input'
				cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', 
																vertical = 'center', 
																wrap_text = True)
		#wb.save('%s' %(name))
		return wb
	except Exception as ex:
		return ('createJournal failed with: %s' %ex)

def clearJournal(journal):
	try:
		wbToCopy = openpyxl.load_workbook(journal)
		for sheet in wbToCopy.worksheets:
			for row in sheet.iter_rows(min_row = 2):
				for cell in row:
					cell.value = None
					cell.style = 'Normal'
		wbToCopy.save(journal)
		return None
	except Exception as ex:
		return ('clearJournal failed with: %s' %ex)

def compileFile(jList,journal): #by cells
	try:
		wbToCopy = createJournal() # openpyxl.load_workbook(journal)#Workbook() #переезжаем в оперативку	
	except Exception as ex:
		return 'CompileJournal failed with %s' %ex
	tempLastRow = {'ДД':2,'ДТЭ':2}
	#[print(item) for item in jList]
	for jItem in jList: 
		wbFromCopy = openpyxl.load_workbook(jItem,read_only=True)
		for sheet in wbFromCopy.sheetnames:
			try:
				subname = sheet.split('_')[0]
				sheetToCopy = wbToCopy[subname] #copy cells to the same sheet in new workbook
				sheetFromCopy = wbFromCopy[sheet]
			except:	continue
			for row in sheetFromCopy.iter_rows(min_row = 2):
				stopFill = False
				isEmptyRow = True
				lastRow = tempLastRow[subname]
				for iCell,cell in enumerate(row):
					if cell.value in ['+','-','*']: stopFill = True
					if (cell.value is None):
						toCopyCell = sheetToCopy.cell(row = lastRow,column = iCell+1, value = ' ')
						if stopFill: toCopyCell.style = 'Normal'
						else: toCopyCell.style = 'Good'
						border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                     						right=openpyxl.styles.Side(style='thin'),
                     						top=openpyxl.styles.Side(style='thin'),
                     						bottom=openpyxl.styles.Side(style='thin'))
						toCopyCell.border = border
					else:
						isEmptyRow = False 
						toCopyCell = sheetToCopy.cell(row = lastRow,column = cell.column, value = cell.value)
						toCopyCell.font = cell.font
						toCopyCell.border = cell.border
						toCopyCell.fill = cell.fill
						toCopyCell.number_format = cell.number_format
						toCopyCell.alignment = cell.alignment
				if isEmptyRow:	continue
				else: tempLastRow[subname]+=1
		wbFromCopy.close()
	for sheet in wbToCopy.worksheets:
		sheet.auto_filter.ref = sheetToCopy.calculate_dimension()
	while True:
		try:
			wbToCopy.save(journal)
			break
		except Exception as ex:
			print ('clearJournal failed with: %s' %ex)
	wbToCopy.close()
	return None # ¯\_(ツ)_/¯
	
def main():
	tStart = dt.datetime.now()
	workDir = whatOSRun(os.getcwd()) #определяем параметры ввода пути до файла
	journalList,journalFile = findFiles(workDir) 
	clearJournal(journalFile)
	compileFile(journalList,journalFile)
	print(dt.datetime.now()- tStart)
	'''при сохранении журнала не устанавливается флаг группового доступа
	tStart = dt.datetime.now()
	stopSwitch = False
	while not(stopSwitch):
		clearJournal(journalFile)
		compileFile(journalList,journalFile)
		print('waiting for 30 minutes')
		tt.sleep(1800)
		stopSwitch = ((dt.datetime.now().hour - tStart.hour) > 9)
	'''
	sys.exit(0)
	
if __name__ == '__main__':
	main()