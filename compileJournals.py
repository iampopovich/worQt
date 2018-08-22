#!/usr/bin/pythom
#-*-coding: utf-8-*-
#import math
import openpyxl
import time as tt
import datetime as dt
import re
import os
#from os import path
import sys
import lib_headers as hh
'''
def findFiles(cwd):
def createJournal(cwd):
def compileFile(jList,config,journal,cwd):
def readWriteConfig(conf,mode,dict = None):
'''
def whatOSRun(path):
	if os.name.endswith('nt'):
		return '%s\\' %path
	else: return '%s/' %path

def findFiles(cwd):
	currentYear = dt.datetime.now().year
	files = [os.path.join(cwd, f) for f in os.listdir(cwd) if os.path.isfile(os.path.join(cwd, f))]
	try:	#этих файлов может и не быть , на всякий случай расставь экспы
		cJournalName = '%sСводный_журнал_ДТЭ_%s.xlsx' %(cwd,currentYear)
		if cJournalName in files: files.pop(files.index(cJournalName))
		else: createJournal(cJournalName)
		xlFiles = [item for item in files if not(re.search(r'.*\sДТЭ\s.*%s.*xlsm' %currentYear, item) is None) and not('$' in item)] # дичь дважды - экранировать доллар научись 
		return xlFiles,cJournalName
	except Exception as ex:
		print('findFiles failed with: %s' %ex)
		return None, None

def createJournal(name):
	try:
		wb = openpyxl.Workbook()
		wb.remove(wb.active)
		for sheetName in ['ДД','ДТЭ']:
			wb.create_sheet(sheetName)
			activeSheet = wb[sheetName]
			activeSheet.append(hh.getHeaders(sheetName))
			for item in activeSheet.columns:
				activeSheet.column_dimensions['%s' %item[0].column].width = 16.0
			#activeSheet.auto_filter.ref = 
			for cell in list(activeSheet.rows)[0]:
				horizontal = 'center'
				vertical = 'center'
				cell.style = 'Input'
				cell.alignment = openpyxl.styles.Alignment(horizontal=horizontal, 
																vertical=vertical, 
																wrap_text=True)
		wb.save('%s' %(name))
		return None
	except Exception as ex:
		return ('createJournal failed with: %s' %ex)

def clearJournal(journal):
	wbToCopy = openpyxl.load_workbook(journal)
	for sheet in wbToCopy.sheetnames:
		cSheet = wbToCopy[sheet]
		for row in cSheet.iter_rows(min_row = 2): #check specs for iter_rows
			for cell in row:
				cell.value = None
	wbToCopy.save(journal)

def compileFile(jList,journal): #by cells
	try:
		wbToCopy = openpyxl.load_workbook(journal)
	except Exception as ex:
		return 'CompileJournal failed with %s' %ex
	tempLastRow = {'ДД':2,'ДТЭ':2} 
	for jItem in jList: 
		wbFromCopy = openpyxl.load_workbook(jItem,read_only=True)
		for sheetFromCopy in wbFromCopy.sheetnames:
			try:
				subname = sheetFromCopy.split('_')[0]
				sheetToCopy = wbToCopy[subname] #copy cells to the same sheet in new workbook
				sheet = wbFromCopy[sheetFromCopy]
			except:	continue
			for row in sheet.iter_rows(min_row = 2):
				isEmptyRow = True
				lastRow = tempLastRow[subname]
				for iCell,cell in enumerate(row):
					if (cell.value is None):
						newCell = sheetToCopy.cell(row = lastRow,column = iCell+1, value = ' ')
						newCell.style = 'Good'
						border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                     						right=openpyxl.styles.Side(style='thin'),
                     						top=openpyxl.styles.Side(style='thin'),
                     						bottom=openpyxl.styles.Side(style='thin'))
						newCell.border = border
					else:
						isEmptyRow = False 
						newCell = sheetToCopy.cell(row = lastRow,column = cell.column, value = cell.value)
						newCell.font = cell.font
						newCell.border = cell.border
						newCell.fill = cell.fill
						newCell.number_format = cell.number_format
						newCell.alignment = cell.alignment
				if isEmptyRow:	continue
				else: tempLastRow[subname]+=1
			sheetToCopy.auto_filter.ref = sheetToCopy.calculate_dimension()
		wbFromCopy.close()
		del wbFromCopy, sheetFromCopy 
		wbToCopy.save(journal)
	wbToCopy.close()
	return None # ¯\_(ツ)_/¯
	
def main():
	tStart = dt.datetime.now()
	workDir = whatOSRun(os.getcwd()) #определяем параметры ввода пути до файла
	stopSwitch = False
	journalList,journalFile = findFiles(workDir) 
	while not(stopSwitch):
		print('start')
		#clearJournal(journalFile)
		compileFile(journalList,journalFile)
		print('waiting for 30 minutes')
		tt.sleep(1800) 
		stopSwitch = ((dt.datetime.now().hour - tStart.hour) > 9)
	sys.exit(0)
	
if __name__ == '__main__':
	main()
